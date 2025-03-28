import requests
from openpyxl import Workbook

# Настройки Prometheus
PROMETHEUS_URL = "http://10.138.55.242:9090/api/v1/query"
INSTANCE_FILTER = 'dc2.*'

# Формулы метрик
queries = {
    "CPU_MAX": f'100 * max by (instance) (1 - rate(node_cpu_seconds_total{{mode="idle", instance=~"{INSTANCE_FILTER}"}}[30d]))',
    "CPU_AVG": f'100 * avg by (instance) (1 - rate(node_cpu_seconds_total{{mode="idle", instance=~"{INSTANCE_FILTER}"}}[30d]))',
    "MEM_MAX": f'((max_over_time(node_memory_MemTotal_bytes{{instance=~"{INSTANCE_FILTER}", job="node_exporter_centos"}}[30d]) - max_over_time(node_memory_MemAvailable_bytes{{instance=~"{INSTANCE_FILTER}", job="node_exporter_centos"}}[30d])) / max_over_time(node_memory_MemTotal_bytes{{instance=~"{INSTANCE_FILTER}", job="node_exporter_centos"}}[30d])) * 100',
    "MEM_AVG": f'((avg_over_time(node_memory_MemTotal_bytes{{instance=~"{INSTANCE_FILTER}", job="node_exporter_centos"}}[30d]) - avg_over_time(node_memory_MemAvailable_bytes{{instance=~"{INSTANCE_FILTER}", job="node_exporter_centos"}}[30d])) / avg_over_time(node_memory_MemTotal_bytes{{instance=~"{INSTANCE_FILTER}", job="node_exporter_centos"}}[30d])) * 100',
    "DISK_UTIL": f'(100 * (1 - (sum by (instance) (node_filesystem_avail_bytes{{instance=~"{INSTANCE_FILTER}", mountpoint!~"/boot|/dev"}}) / sum by (instance) (node_filesystem_size_bytes{{instance=~"{INSTANCE_FILTER}", mountpoint!~"/boot|/dev"}}))))'
}

# Функция для получения данных из Prometheus
def get_prometheus_data(query):
    response = requests.get(PROMETHEUS_URL, params={"query": query})
    if response.status_code == 200:
        data = response.json()
        if data['status'] == 'success' and data['data']['result']:
            return {item['metric']['instance']: float(item['value'][1]) for item in data['data']['result']}
    return {}

# Получение данных для всех метрик
results = {}
for name, query in queries.items():
    results[name] = get_prometheus_data(query)

# Создание Excel-файла
wb = Workbook()
ws = wb.active

# Заголовки
ws.cell(row=1, column=1, value="Instance")
ws.cell(row=1, column=2, value="CPU_MAX (%)")
ws.cell(row=1, column=3, value="CPU_AVG (%)")
ws.cell(row=1, column=4, value="MEM_MAX (%)")
ws.cell(row=1, column=5, value="MEM_AVG (%)")
ws.cell(row=1, column=6, value="DISK_UTIL (%)")

# Заполнение данных
instances = set(results["CPU_MAX"].keys()).union(
    results["CPU_AVG"].keys(),
    results["MEM_MAX"].keys(),
    results["MEM_AVG"].keys(),
    results["DISK_UTIL"].keys()
)
row_index = 2
for instance in sorted(instances):
    ws.cell(row=row_index, column=1, value=instance)
    ws.cell(row=row_index, column=2, value=results["CPU_MAX"].get(instance, "N/A"))
    ws.cell(row=row_index, column=3, value=results["CPU_AVG"].get(instance, "N/A"))
    ws.cell(row=row_index, column=4, value=results["MEM_MAX"].get(instance, "N/A"))
    ws.cell(row=row_index, column=5, value=results["MEM_AVG"].get(instance, "N/A"))
    ws.cell(row=row_index, column=6, value=results["DISK_UTIL"].get(instance, "N/A"))
    row_index += 1

# Сохранение файла
wb.save("prometheus_metrics_30_days.xlsx")
print("Данные сохранены в prometheus_metrics_30_days.xlsx")
